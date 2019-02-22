from .models import RepricingOMXData

import csv
import pandas as pd
from pandas import read_csv
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.styles.borders import Border, Side

def convert_queryset_csv(query,file_name):

    df = pd.DataFrame(list(query.values('vendorCode','bfdItemCode',
                     'vendor_UPC','vendor_ProductName',
                     'currWholesalePrice_AD','newWholesalePrice_AD',
                     'currWholesalePrice_BD','newWholesalePrice_BD',
                     'currDropShipPrice_AD','newDropShipPrice_AD',
                     'currDropShipPrice_BD','newDropShipPrice_BD',
                     'dropShipFee','mapPrice','fullRetailPrice_MSRP',
                     'optionOneValue','optionOneSurcharge',
                     'optionTwoValue','optionTwoSurcharge',
                     'optionThreeValue','optionThreeSurcharge')))


    df.rename(columns={'bfdItemCode':'Item Code',"vendorCode":'Manufacturer Model Number',
            'vendor_UPC':'UPC Code','vendor_ProductName':'Product Name',
            'currWholesalePrice_AD':'Current Wholesale (Stocking) Price AFTER Discount',
            'newWholesalePrice_AD':'New Wholesale (Stocking) Price AFTER Discount',
            'currWholesalePrice_BD':'Current Wholesale (Stocking) Price BEFORE Discount',
            'newWholesalePrice_BD':'New Wholesale (Stocking) Price BEFORE Discount',
            'currDropShipPrice_AD':'Current Drop Ship Price AFTER Discount',
            'newDropShipPrice_AD':'New Drop Ship Price AFTER Discount',
            'currDropShipPrice_BD':'Current Drop Ship Price BEFORE Discount',
            'newDropShipPrice_BD':'New Drop Ship Price BEFORE Discount',
            'dropShipFee':'Drop Ship Fee (optional)', 'mapPrice':'MAP Price (Sale Price)','fullRetailPrice_MSRP':'Full Retail Price (MSRP)',
            "optionOneValue":"Option 1 Value","optionOneSurcharge":"Option 1 Surcharge",
            "optionTwoValue":"Option 2 Value","optionTwoSurcharge":"Option 2 Surcharge",
            "optionThreeValue":"Option 3 Value","optionThreeSurcharge":"Option 3 Surcharge"},inplace=True)

    headers=['Item Code','Manufacturer Model Number','UPC Code','Product Name',
            'Current Wholesale (Stocking) Price AFTER Discount',
            'New Wholesale (Stocking) Price AFTER Discount',
            'Current Wholesale (Stocking) Price BEFORE Discount',
            'New Wholesale (Stocking) Price BEFORE Discount',
            'Current Drop Ship Price AFTER Discount',
            'New Drop Ship Price AFTER Discount',
            'Current Drop Ship Price BEFORE Discount',
            'New Drop Ship Price BEFORE Discount',
            'Drop Ship Fee (optional)','MAP Price (Sale Price)','Full Retail Price (MSRP)',
            'Option 1 Value',"Option 1 Surcharge",
            "Option 2 Value","Option 2 Surcharge",
            "Option 3 Value","Option 3 Surcharge"]

    df.to_csv(file_name, sep=',', encoding='utf-8',index=False, header=True,
            columns=headers)

    return file_name

def format_excel_file(filepath, xl_file_name):

    dfUpdated = pd.read_csv(filepath, encoding='latin-1')

    greenFill = PatternFill(start_color='5cd65c',
                            end_color='5cd65c',
                            fill_type='solid')
    yellowFill = PatternFill(start_color='ffff4d',
                            end_color='ffff4d',
                            fill_type='solid')
    blueFill = PatternFill(start_color='80bfff',
                            end_color='80bfff',
                            fill_type='solid')
    redFill = PatternFill(start_color='ff704d',
                            end_color='ff704d',
                            fill_type='solid')
    right_border = Border(right=Side(style='thin'))

    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendor Repricing'

    ws['A1'] = 'Fullfillment Information'
    ws['P1'] = 'Options'
    ws.merge_cells('A1:O1')
    ws.merge_cells('P1:U1')

    for row in dataframe_to_rows(dfUpdated, index=False, header=True):
        ws.append(row)

    for column_cells in ws.columns:
        ws.column_dimensions[column_cells[0].column].width = 20

    for row in ws.iter_rows():
        if row[0].value == "Fullfillment Information":
            for cell in row:
                if cell.value == 'Fullfillment Information':
                    ws[str(cell.coordinate)].fill = redFill
                    ws[str(cell.coordinate)].border = thin_border
                    ws[str(cell.coordinate)].alignment = Alignment(horizontal='center')
                if cell.value == 'Options':
                    ws[str(cell.coordinate)].fill = blueFill
                    ws[str(cell.coordinate)].border = thin_border
                    ws[str(cell.coordinate)].alignment = Alignment(horizontal='center')
        elif row[0].value == 'Item Code':
            for cell in row:
                if cell.value == 'New Wholesale (Stocking) Price AFTER Discount' or cell.value == 'New Wholesale (Stocking) Price BEFORE Discount' or cell.value == 'New Drop Ship Price AFTER Discount' or cell.value == 'New Drop Ship Price BEFORE Discount':
                    ws[str(cell.coordinate)].fill = greenFill
                    ws[str(cell.coordinate)].border = thin_border
                elif (cell.value == 'Option 1 Value' or cell.value == 'Option 1 Surcharge' or
                      cell.value == 'Option 2 Value' or cell.value == 'Option 2 Surcharge' or
                      cell.value == 'Option 3 Value' or cell.value == 'Option 3 Surcharge'):
                    ws[str(cell.coordinate)].fill = blueFill
                    ws[str(cell.coordinate)].border = thin_border
                else:
                    ws[str(cell.coordinate)].fill = yellowFill
                    ws[str(cell.coordinate)].border = thin_border

        else:
            row[3].border = right_border
            row[5].border = right_border
            row[7].border = right_border
            row[9].border = right_border
            row[11].border = right_border
    wb.save(xl_file_name)
    return xl_file_name

def as_text(value):
    if value is None:
        return ""
    return str(value)
