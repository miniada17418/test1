import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell


import datetime
from datetime import date, timedelta
import pandas as pd
from pandas import read_csv
import csv, codecs
import os

from .models import UploadFile, VendorNames, OMXOrderHistory, QBImportFile, NeedToAddVendorToDatabase

def handle_uploaded_file(f):
    #converts the file name to include today's date
    time = datetime.datetime.today().strftime('%d-%m-%y')
    fileLocation = 'media/qbreports/uploads/qbUpload{}.csv'.format(time)
    with open(fileLocation, 'wb') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    #function that formats the file for uploading into the QBImportFile table
    # formatfile(fileLocation)
    # uploadToDatabase
    return fileLocation

def formatfile(f):

    headers = ["Janet's Notes","CustomID", "Vendor Name",
               "Invoice  Date", "Invoice Number", "PO Num",
              "Invoice Total", "Drop Ship Fee", "Shipping Fee", "OMX PO",
              "OMX Total", "Total    Difference", "Discrepancy", 'Memo'
              ]
    # df = pd.DataFrame(list(VendorNames.objects.all().values()))
    didntpost = []
    updatedQBfile = 'media/qbreports/uploads/updatedQBfile.csv'
    with open(updatedQBfile, 'w') as csvWriter,  open(f,'rt', encoding='utf8') as csvfile:
        writer = csv.DictWriter(csvWriter, fieldnames=headers, extrasaction='ignore', lineterminator='\n')
        readCSV = csv.DictReader(csvfile, delimiter=',')
        writer.writeheader()
        for row in readCSV:
            if row["OMX PO"] == None or row["OMX PO"] == "":
                if row["OMX Total"] == "" or row["OMX Total"] == None:
                    writer.writerow(row)
                else:
                    row['OMX Total'] = float(row["OMX Total"].replace(',',''))
                    if row['Invoice Total']:
                        row['Invoice Total'] = float(row['Invoice Total'].replace(',',''))
                    writer.writerow(row)
            else:

                if row["OMX Total"] == "" or row["OMX Total"] == None:
                    writer.writerow(row)
                else:
                    row['OMX Total'] = float(row["OMX Total"].replace(',',''))
                    if row['Invoice Total']:
                        row['Invoice Total'] = float(row['Invoice Total'].replace(',',''))
                    writer.writerow(row)

    # if vendor_name_not_in_database:
    #     need_to_add_vendor_to_database_delete_everything()
    #     for vendor in vendor_name_not_in_database:
    #         try:
    #             _, created = NeedToAddVendorToDatabase.objects.get_or_create(
    #             verndorname = vendor
    #             )
    #         except:
    #             continue
    #
    #     return updatedQBfile, True
    # else:
        #Uplads the file to the QBImportFile table
        return updatedQBfile, False

def uploadToDatabase(f):
    print('started QBImport DB update')
    with open(f, 'r') as post:
        reader = csv.reader(post)
        reader.__next__()
        for row in reader:
            if '/' in row[3]:
                idate = row[3].split('/')
                newIdate = idate[2]+'-'+idate[0]+'-'+idate[1]
                _, created = QBImportFile.objects.get_or_create(
                notes = row[0],
                customID = notNull(row[1]),
                vendorName = notNull(row[2]),
                invoiceDate = newIdate,
                invoiceNumber = notNull(row[4]),
                poNumb = notNull(row[5]),
                invocieTotal = notNull(row[6]),
                dropShipFee = notNull(row[7]),
                shippingFee = notNull(row[8]),
                omxPO = notNull(row[9]),
                omxTotal = notNull(row[10]),
                totalDifference = notNull(row[11]),
                discrepancy = row[12],
                memo = notNull(row[13])
                )

            else:
                _, created = QBImportFile.objects.get_or_create(
                notes = row[0],
                customID = notNull(row[1]),
                vendorName = notNull(row[2]),
                invoiceDate = from_excel_ordinal(int(row[3])).strftime("%Y-%m-%d"),
                invoiceNumber = notNull(row[4]),
                poNumb = notNull(row[5]),
                invocieTotal = notNull(row[6]),
                dropShipFee = notNull(row[7]),
                shippingFee = notNull(row[8]),
                omxPO = notNull(row[9]),
                omxTotal = notNull(row[10]),
                totalDifference = notNull(row[11]),
                discrepancy = row[12],
                memo = notNull(row[13])
                )
    print("finished QBImportFIle DB Update")

def join_Database():

    xlsx_to_formated_xlsx = 'media/qbreports/uploads/xlsx_to_formated_xlsx.xlsx'
    downlaod_file_name = 'media/qbreports/uploads/custom_qb_report.xlsx'

    joined_table_file_name = pandas_df_join()

    csv_to_xlsx, google_sheet = finial_file_creation(joined_table_file_name)


    dfUpdated = pd.read_csv(csv_to_xlsx, encoding='latin-1')
    dfgdoc = pd.read_csv(google_sheet, encoding='latin-1')

    print("creating xlsx files")

    wb = Workbook()
    ws = wb.active
    ws.title = 'Matching_Updated'
    ws2 = wb.create_sheet()
    ws2.title = 'Vendor Price Check Gdoc'

    bold = NamedStyle(name='bold')
    bold.font = Font(bold=True)
    wb.add_named_style(bold)

    cell = WriteOnlyCell(ws)
    cell.style = 'bold'

    greenFill = PatternFill(start_color='32CD32',
                            end_color='32CD32',
                            fill_type='solid')


    for row in dataframe_to_rows(dfUpdated,index=False, header=True):
        ws.append(row)

    for row in dataframe_to_rows(dfgdoc, index=False, header=True):
        ws2.append(row)

    wb.save(xlsx_to_formated_xlsx)

    wb = load_workbook(filename = xlsx_to_formated_xlsx)
    ws = wb.active

    for row in ws.iter_rows():
        if row[1].value == "Vendor Name":
            for cell in row:
                ws[str(cell.coordinate)].font = Font(bold=True)
        if row[1].value == None:
            for cell in row:
                if cell.value != None:
                    ws[str(cell.coordinate)].fill = greenFill
        if row[1].value != None:
            for cell in row:
                ws[str(cell.coordinate)].font = Font(bold=True)

    wb.save(downlaod_file_name)
    print("created xlsx files")

    qb_Import_File_delete_everything()

    return downlaod_file_name

def notNull(x):
    if x == None or x == '':
        return 0
    else:
        return x

#converts excel date number into accurate datetime format
def from_excel_ordinal(ordinal, _epoch=date(1900, 1, 1)):
    """
    formates excels unformated date value into the correct format for adding The
    value to the Table.
    """
    if ordinal > 59:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    return _epoch + timedelta(days=ordinal - 1)  # epoch is day 1

def qb_Import_File_delete_everything():
    """
    Deletes all data from the QBImportFile Table
    """
    QBImportFile.objects.all().delete()

def need_to_add_vendor_to_database_delete_everything():
    """
    Deletes all data from the QBImportFile Table
    """
    NeedToAddVendorToDatabase.objects.all().delete()

def pandas_df_join():
    print('started merging Pandas DF')
    joined_table_file_name = 'media/qbreports/uploads/qb_omx_joined_tables.csv'
    df = pd.DataFrame(list(QBImportFile.objects.all().values('notes','vendorName',
                                                             'invoiceDate','invoiceNumber',
                                                             'invocieTotal','dropShipFee',
                                                             'shippingFee', 'omxPO',
                                                             'omxTotal','totalDifference',
                                                             'discrepancy','memo',
                                                             'poNumb')))

    df = df[['notes','vendorName','invoiceDate','invoiceNumber',
             'invocieTotal','dropShipFee','shippingFee', 'omxPO','omxTotal','totalDifference',
             'discrepancy','memo','poNumb']]

    df2 = pd.DataFrame(list(OMXOrderHistory.objects.all().values('orderNumber', 'itemCode',
                                                            'description','dropShipCOG',
                                                            'qty','total')))
    df2 = df2[['orderNumber', 'itemCode', 'description','dropShipCOG','qty','total']]

    df3 = pd.merge(df, df2, left_on='poNumb', right_on='orderNumber', how='left')
    print("Finished Pandas DB Merge")

    df3.to_csv(joined_table_file_name, sep=',', encoding='utf-8')
    print("created Merged .csv file")
    return joined_table_file_name

def finial_file_creation(x):
    print("Creating Janet's files")
    csv_file_to_xlsx = 'media/qbreports/uploads/csv_file_to_xlsx.csv'
    google_doc_sheet = 'media/qbreports/uploads/google_doc_details.csv'

    writeFileList = []
    writeFileList.append(["Notes",'Vendor Name','Invoice Date', 'Invoice Number',
                          'PO Num', 'Invoice Total', 'Drop Ship Fee',
                          'Shipping Fee', 'OMX PO', 'OMX Total',
                          'Total Difference', 'Discrepancy','Memo'])

    itemHeader = ["","","","","","",'Code','Item Name', 'QTY',
                  'OMX Unit Price (Drop Ship)', 'Drop Ship Total',"" ,""
                  ]


    google_doc_headers = ['Vendor Name','Entry Date', 'Invoice Date', 'Invoice Number',
                          'OMX Order Number', 'Quantity', "Item Code",
                          'OMX Price', 'Invoice Price'
                          ]

    moreThanOneItem = []

    with open(x, 'rt', encoding='utf8') as f, open(csv_file_to_xlsx,'w', newline='',encoding='utf8') as writeCSV, open(google_doc_sheet, 'w', newline='',encoding='utf8') as gdocCSV:
        write = csv.writer(writeCSV, delimiter=',')
        gwrite = csv.writer(gdocCSV, delimiter=',')
        readCSV = csv.DictReader(f, delimiter=',' )
        fileList = []
        newList = []
        google_doc_list = []
        google_doc_list.append(google_doc_headers)
        for row in readCSV:
            fileList.append(row)


        for i in range(len(fileList)):
            google_doc_format = [fileList[i]['vendorName'], datetime.datetime.now().date() ,fileList[i]['invoiceDate'],
                                fileList[i]['invoiceNumber'],fileList[i]['omxPO'],
                                fileList[i]['qty'], fileList[i]['itemCode'],
                                fileList[i]['dropShipCOG'],fileList[i]['invocieTotal']
                                ]
            invoiceDetails = [fileList[i]['notes'],fileList[i]['vendorName'], fileList[i]['invoiceDate'],
                             fileList[i]['invoiceNumber'],fileList[i]['poNumb'],
                             fileList[i]['invocieTotal'],
                             fileList[i]['dropShipFee'], fileList[i]['shippingFee'],
                             fileList[i]['omxPO'], fileList[i]['omxTotal'],
                             fileList[i]['totalDifference'], fileList[i]['discrepancy'],
                             fileList[i]['memo']
                             ]


            productDetails = ["","","",fileList[i]['invoiceNumber'],fileList[i]['poNumb'],"",
                              fileList[i]['itemCode'],fileList[i]['description'],
                              fileList[i]['qty'], fileList[i]['dropShipCOG'],
                              fileList[i]['total'],"" ,""]
            try:
                if fileList[i]['orderNumber'] in moreThanOneItem and fileList[i]['invoiceNumber'] == fileList[i-1]['invoiceNumber']:
                    writeFileList.append(productDetails)
                    google_doc_list.append(google_doc_format)

                # elif fileList[i]['discrepancy'] == None or fileList[i]['discrepancy'] == 'Duplicate - NOT UPLOADED' or fileList[i]['discrepancy'] == '':
                #     if fileList[i]['omxPO'] == fileList[i+1]['omxPO']:
                #         continue
                #     else:
                #         writeFileList.append(invoiceDetails)

                elif fileList[i]['orderNumber'] == '' or fileList[i]['orderNumber'] == None:
                    writeFileList.append(invoiceDetails)

                elif fileList[i]['orderNumber'] == fileList[i+1]['orderNumber'] and fileList[i]['invoiceNumber'] == fileList[i+1]['invoiceNumber']:
                    moreThanOneItem.append(fileList[i]['orderNumber'])
                    writeFileList.append(invoiceDetails)
                    writeFileList.append(itemHeader)
                    writeFileList.append(productDetails)
                    google_doc_list.append(google_doc_format)
                else:
                    writeFileList.append(invoiceDetails)
                    writeFileList.append(itemHeader)
                    writeFileList.append(productDetails)
                    google_doc_list.append(google_doc_format)
            except IndexError:
                print("INDEX ERROR: " + fileList[i]["orderNumber"] + " " + fileList[i]["itemCode"])
                print(newList.append(invoiceDetails))
                writeFileList.append(invoiceDetails)
                writeFileList.append(itemHeader)
                writeFileList.append(productDetails)
                google_doc_list.append(google_doc_format)

        for x in writeFileList:
            write.writerow(x)

        for x in google_doc_list:
            gwrite.writerow(x)
    return csv_file_to_xlsx , google_doc_sheet
