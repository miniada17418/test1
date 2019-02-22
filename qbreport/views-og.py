from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.generic import (TemplateView, ListView,
                                    DetailView, CreateView)
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.conf import settings
from django.template import loader
from wsgiref.util import FileWrapper

# Create your views here.
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell
import datetime
from datetime import date, timedelta
import pandas as pd
from pandas import read_csv
import csv, codecs
import os


from .forms import DocumenUpload
from .models import UploadFile, VendorNames, OMXOrderHistory, QBImportFile, NeedToAddVendorToDatabase


#View that pulls in the file
def modelFormUpload(request):
    if request.method == 'POST':
        form = DocumenUpload(request.POST, request.FILES)
        if form.is_valid():
            #Function that saves the file with a unique name
            #formats the file to match the QBImportFile Table
            #Then uploads the file data to the database
            UploadedFilePath = handle_uploaded_file(request.FILES['orderReport'])
            QBFile, vendorissue = formatfile(UploadedFilePath)

            if vendorissue:
                return redirect('qb:vendor_not_in_db')
            else:
                uploadToDatabase(QBFile)

            finialFormatedFile = join_Database()

            obj = form.save(commit=False)
            obj.userName = request.user
            obj.formatedFile = finialFormatedFile
            obj.save()

            #saves the user and email address to the UploadFile table
            form.save()

            NewFileName = finialFormatedFile.split('/')

            return redirect('qb:download_file', file_name=NewFileName[-1])

    else:
        form = DocumenUpload()
    return render(request, 'qbreport/uploadfile.html', {
        'form': form
    })

def handle_uploaded_file(f):
    #converts the file name to include today's date
    time = datetime.datetime.today().strftime('%d-%m-%y')
    fileLocation = 'media/qbreports/uploads/qbUpload{}.csv'.format(time)
    with open(fileLocation, 'wb') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    #function that formats the file for uploading into the QBImportFile table
    # formatfile(fileLocation)
    # uploadToDatabase
    return fileLocation


def formatfile(f):

    headers = ["Janet's Notes","CustomID", "Vendor Name",
               "Invoice  Date", "Invoice Number", "PO Num",
              "Invoice Total", "Drop Ship Fee", "Shipping Fee", "OMX PO",
              "OMX Total", "Total    Difference", "Discrepancy", 'Memo'
              ]
    df = pd.DataFrame(list(VendorNames.objects.all().values()))
    didntpost = []
    vendor_name_not_in_database = []
    updatedQBfile = 'media/qbreports/uploads/updatedQBfile.csv'
    with open(updatedQBfile, 'w') as csvWriter,  open(f,'rt', encoding='utf8') as csvfile:
        writer = csv.DictWriter(csvWriter, fieldnames=headers, extrasaction='ignore', lineterminator='\n')
        readCSV = csv.DictReader(csvfile, delimiter=',')
        writer.writeheader()
        for row in readCSV:
            sameVendor = df.loc[df['qbVendorName'] == row["Vendor Name"]]
            try:
                #sets the vendor name in the vendorNameUpdatedFile
                row["Vendor Name"] = sameVendor['omxVendorName'].values[0]
            except:
                vendor_name_not_in_database.append(row["Vendor Name"])

            if row["OMX PO"] == None or row["OMX PO"] == "":
                if row["OMX Total"] == "" or row["OMX Total"] == None:
                    writer.writerow(row)
                else:
                    row['OMX Total'] = float(row["OMX Total"].replace(',',''))
                    if row['Invoice Total']:
                        row['Invoice Total'] = float(row['Invoice Total'].replace(',',''))
                    writer.writerow(row)
            else:
                #CustomID is the value used to find which products in an order
                #are assigned to which vendor
                row["CustomID"] = row["Vendor Name"] + row["OMX PO"]
                if row["OMX Total"] == "" or row["OMX Total"] == None:
                    writer.writerow(row)
                else:
                    row['OMX Total'] = float(row["OMX Total"].replace(',',''))
                    if row['Invoice Total']:
                        row['Invoice Total'] = float(row['Invoice Total'].replace(',',''))
                    writer.writerow(row)

    if vendor_name_not_in_database:
        need_to_add_vendor_to_database_delete_everything()
        for vendor in vendor_name_not_in_database:
            try:
                _, created = NeedToAddVendorToDatabase.objects.get_or_create(
                verndorname = vendor
                )
            except:
                continue

        return updatedQBfile, True
    else:
        #Uplads the file to the QBImportFile table
        return updatedQBfile, False

def uploadToDatabase(f):
    with open(f, 'r') as post:
        reader = csv.reader(post)
        reader.__next__()
        for row in reader:
            if '/' in row[3]:
                idate = row[3].split('/')
                newIdate = idate[2]+'-'+idate[0]+'-'+idate[1]
                _, created = QBImportFile.objects.get_or_create(
                notes = row[0],
                customID = notNull(row[1]),
                vendorName = notNull(row[2]),
                invoiceDate = newIdate,
                invoiceNumber = notNull(row[4]),
                poNumb = notNull(row[5]),
                invocieTotal = notNull(row[6]),
                dropShipFee = notNull(row[7]),
                shippingFee = notNull(row[8]),
                omxPO = notNull(row[9]),
                omxTotal = notNull(row[10]),
                totalDifference = notNull(row[11]),
                discrepancy = row[12],
                memo = notNull(row[13])
                )

            else:
                _, created = QBImportFile.objects.get_or_create(
                notes = row[0],
                customID = notNull(row[1]),
                vendorName = notNull(row[2]),
                invoiceDate = from_excel_ordinal(int(row[3])).strftime("%Y-%m-%d"),
                invoiceNumber = notNull(row[4]),
                poNumb = notNull(row[5]),
                invocieTotal = notNull(row[6]),
                dropShipFee = notNull(row[7]),
                shippingFee = notNull(row[8]),
                omxPO = notNull(row[9]),
                omxTotal = notNull(row[10]),
                totalDifference = notNull(row[11]),
                discrepancy = row[12],
                memo = notNull(row[13])
                )

def join_Database():
    joined_table_file_name = 'media/qbreports/uploads/qb_omx_joined_tables.csv'
    csv_file_to_xlsx = 'media/qbreports/uploads/csv_file_to_xlsx.csv'
    xlsx_to_formated_xlsx = 'media/qbreports/uploads/xlsx_to_formated_xlsx.xlsx'
    downlaod_file_name = 'media/qbreports/uploads/custom_qb_report.xlsx'
    google_doc_sheet = 'media/qbreports/uploads/google_doc_details.csv'


    df = pd.DataFrame(list(QBImportFile.objects.all().values('notes','customID','vendorName',
                                                             'invoiceDate','invoiceNumber',
                                                             'poNumb','invocieTotal','dropShipFee',
                                                             'shippingFee', 'omxPO',
                                                             'omxTotal','totalDifference',
                                                             'discrepancy','memo')))

    df = df[['notes','customID','vendorName','invoiceDate','invoiceNumber',
             'poNumb','invocieTotal','dropShipFee','shippingFee', 'omxPO','omxTotal','totalDifference',
             'discrepancy','memo']]

    df2 = pd.DataFrame(list(OMXOrderHistory.objects.all().values('customID', 'itemCode',
                                                            'description','dropShipCOG',
                                                            'qty','total')))
    df2 = df2[['customID', 'itemCode', 'description','dropShipCOG','qty','total']]

    df3 = pd.merge(df, df2, on='customID', how='left')

    df3.to_csv(joined_table_file_name, sep=',', encoding='utf-8')

    writeFileList = []
    writeFileList.append(["Notes",'Vendor Name','Invoice Date', 'Invoice Number',
                          'PO Num', 'Invoice Total', 'Drop Ship Fee',
                          'Shipping Fee', 'OMX PO', 'OMX Total',
                          'Total Difference', 'Discrepancy','Memo'])

    itemHeader = ["","","","","","",'Code','Item Name', 'QTY',
                  'OMX Unit Price (Drop Ship)', 'Drop Ship Total',"" ,""
                  ]
    moreThanOneItem = []

    google_doc_headers = ['Vendor Name', 'Invoice Date', 'Invoice Number',
                          'OMX Order Number', 'Quantity', "Item Code",
                          'OMX Price', 'Invoice Price'
                          ]

    with open(joined_table_file_name, 'rt', encoding='utf8') as f, open(csv_file_to_xlsx,'w', newline='',encoding='utf8') as writeCSV, open(google_doc_sheet, 'w', newline='',encoding='utf8') as gdocCSV:
        write = csv.writer(writeCSV, delimiter=',')
        gwrite = csv.writer(gdocCSV, delimiter=',')
        readCSV = csv.DictReader(f, delimiter=',' )
        fileList = []
        newList = []
        google_doc_list = []
        google_doc_list.append(google_doc_headers)
        for row in readCSV:
            fileList.append(row)

        for i in range(len(fileList)):
            google_doc_format = [fileList[i]['vendorName'],fileList[i]['invoiceDate'],
                                fileList[i]['invoiceNumber'],fileList[i]['omxPO'],
                                fileList[i]['qty'], fileList[i]['itemCode'],
                                fileList[i]['dropShipCOG'],fileList[i]['invocieTotal']
                                ]
            invoiceDetails = [fileList[i]['notes'],fileList[i]['vendorName'], fileList[i]['invoiceDate'],
                             fileList[i]['invoiceNumber'],fileList[i]['poNumb'],
                             fileList[i]['invocieTotal'],
                             fileList[i]['dropShipFee'], fileList[i]['shippingFee'],
                             fileList[i]['omxPO'], fileList[i]['omxTotal'],
                             fileList[i]['totalDifference'], fileList[i]['discrepancy'],
                             fileList[i]['memo']
                             ]


            productDetails = ["","","",fileList[i]['invoiceNumber'],fileList[i]['poNumb'],"",
                              fileList[i]['itemCode'],fileList[i]['description'],
                              fileList[i]['qty'], fileList[i]['dropShipCOG'],
                              fileList[i]['total'],"" ,""]
            try:
                if fileList[i]['omxPO'] in moreThanOneItem and fileList[i]['invoiceNumber'] == fileList[i-1]['invoiceNumber']:
                    writeFileList.append(productDetails)
                    google_doc_list.append(google_doc_format)

                elif fileList[i]['discrepancy'] == None or fileList[i]['discrepancy'] == 'Duplicate - NOT UPLOADED' or fileList[i]['discrepancy'] == '':
                    if fileList[i]['omxPO'] == fileList[i+1]['omxPO']:
                        continue
                    else:
                        writeFileList.append(invoiceDetails)

                elif fileList[i]['omxPO'] == '0':
                    writeFileList.append(invoiceDetails)

                elif fileList[i]['omxPO'] == fileList[i+1]['omxPO'] and fileList[i]['invoiceNumber'] == fileList[i+1]['invoiceNumber']:
                    moreThanOneItem.append(fileList[i]['omxPO'])
                    writeFileList.append(invoiceDetails)
                    writeFileList.append(itemHeader)
                    writeFileList.append(productDetails)
                    google_doc_list.append(google_doc_format)
                else:
                    writeFileList.append(invoiceDetails)
                    writeFileList.append(itemHeader)
                    writeFileList.append(productDetails)
                    google_doc_list.append(google_doc_format)
            except IndexError:
                print(newList.append(invoiceDetails))

        for x in writeFileList:
            write.writerow(x)

        for x in google_doc_list:
            gwrite.writerow(x)

    dfUpdated = pd.read_csv(csv_file_to_xlsx, encoding='latin-1')
    dfgdoc = pd.read_csv(google_doc_sheet, encoding='latin-1')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Matching_Updated'
    ws2 = wb.create_sheet()
    ws2.title = 'Vendor Price Check Gdoc'

    bold = NamedStyle(name='bold')
    bold.font = Font(bold=True)
    wb.add_named_style(bold)

    cell = WriteOnlyCell(ws)
    cell.style = 'bold'

    greenFill = PatternFill(start_color='32CD32',
                            end_color='32CD32',
                            fill_type='solid')


    for row in dataframe_to_rows(dfUpdated,index=False, header=True):
        ws.append(row)

    for row in dataframe_to_rows(dfgdoc, index=False, header=True):
        ws2.append(row)

    wb.save(xlsx_to_formated_xlsx)

    wb = load_workbook(filename = xlsx_to_formated_xlsx)
    ws = wb.active

    for row in ws.iter_rows():
        if row[1].value == "Vendor Name":
            for cell in row:
                ws[str(cell.coordinate)].font = Font(bold=True)
        if row[1].value == None:
            for cell in row:
                if cell.value != None:
                    ws[str(cell.coordinate)].fill = greenFill
        if row[1].value != None:
            for cell in row:
                ws[str(cell.coordinate)].font = Font(bold=True)

    wb.save(downlaod_file_name)

    qb_Import_File_delete_everything()

    return downlaod_file_name



def qb_Import_File_delete_everything():
    """
    Deletes all data from the QBImportFile Table
    """
    QBImportFile.objects.all().delete()

def need_to_add_vendor_to_database_delete_everything():
    """
    Deletes all data from the QBImportFile Table
    """
    NeedToAddVendorToDatabase.objects.all().delete()

#converts excel date number into accurate datetime format
def from_excel_ordinal(ordinal, _epoch=date(1900, 1, 1)):
    """
    formates excels unformated date value into the correct format for adding The
    value to the Table.
    """
    if ordinal > 59:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    return _epoch + timedelta(days=ordinal - 1)  # epoch is day 1

def send_file(request,file_name):
    """
    Send a file through Django without loading the whole file into
    memory at once. The FileWrapper will turn the file object into an
    iterator for chunks of 8KB.
    """
    render(request, 'qbreport/download.html')
    file_path = settings.MEDIA_ROOT +'\\qbreports\\uploads\\'+ file_name
    #filename = 'media/qbreports/uploads/custom_qb_report.xlsx' # Select your file here.
    wrapper = FileWrapper(open(file_path, 'rb'))
    response = HttpResponse(wrapper, content_type='application/vnd.ms-excel')
    response['Content-Length'] = os.path.getsize(file_path)
    return response

def download_template(request):
    """
    When the Downloads a template file of the headers need to
    upload a file correctly to the tool
    """
    file_path = settings.MEDIA_ROOT +'\\qbreports\\uploads\\'+ "qb_report_template_file.csv"
    wrapper = FileWrapper(open(file_path, 'rb'))
    response = HttpResponse(wrapper, content_type='text/csv')
    response['Content-Length'] = os.path.getsize(file_path)
    return response

def notNull(x):
    if x == None or x == '':
        return 0
    else:
        return x

def vendor_not_in_datab(request):
    """
    if a vendor name is not in our database the user will be able to views
    the names of the vendors that need to be added to the Database.
    """

    vendors = NeedToAddVendorToDatabase.objects.all().values()

    return render(request, 'qbreport/new-vendor.html', {'vendors':vendors} )
